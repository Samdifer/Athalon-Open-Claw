#!/usr/bin/env python3
"""
Generate Athelon Parts Department Bulk Upload Template (.xlsx)
Multi-sheet Excel workbook covering all parts department data categories.

v2 — Improved formatting, visual hierarchy, print layout, and usability.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime

wb = openpyxl.Workbook()

# ── Style Palette ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

REQUIRED_FONT = Font(name="Calibri", size=9, bold=True, color="CC0000")
REQUIRED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # light red

OPTIONAL_FONT = Font(name="Calibri", size=9, color="666666")
OPTIONAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # light gray

EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="7F8C8D")
EXAMPLE_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # very light yellow

NOTE_FONT = Font(name="Calibri", size=9, color="4A6785")
NOTE_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")  # very light blue

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="0D3B66"),
    right=Side(style="thin", color="0D3B66"),
    top=Side(style="medium", color="0D3B66"),
    bottom=Side(style="medium", color="0D3B66"),
)

# Tab color groups
TAB_PARTS = "2E75B6"       # blue — parts sheets
TAB_WAREHOUSE = "00897B"   # teal — warehouse/location
TAB_TOOLS = "F57C00"       # orange — tools/equipment
TAB_VENDORS = "7B1FA2"     # purple — vendors/PO/cores
TAB_CONFIG = "558B2F"      # green — tags/kits/config
TAB_README = "D32F2F"      # red — instructions


def style_header_row(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    ws.row_dimensions[row].height = 30


def style_indicator_row(ws, num_cols, required_flags, row=2):
    """Row 2: visual distinction between REQUIRED (red tint) and optional (gray tint)."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        is_req = required_flags[col - 1]
        cell.value = "REQUIRED" if is_req else "optional"
        cell.font = REQUIRED_FONT if is_req else OPTIONAL_FONT
        cell.fill = REQUIRED_FILL if is_req else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18


def style_example_row(ws, examples, row=3):
    for col, ex in enumerate(examples, 1):
        cell = ws.cell(row=row, column=col, value=ex)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def style_notes_row(ws, notes, row=4):
    if not notes:
        return
    for col, note in enumerate(notes, 1):
        cell = ws.cell(row=row, column=col)
        if note:
            # If note is very long, put abbreviated text in cell + full text in comment
            if len(str(note)) > 45:
                short = str(note)[:42] + "..."
                cell.value = short
                cell.comment = Comment(str(note), "Athelon Template")
                cell.comment.width = 350
                cell.comment.height = 100
            else:
                cell.value = note
        cell.font = NOTE_FONT
        cell.fill = NOTE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 36


def auto_width(ws, min_width=14, max_width=30):
    """Auto-fit columns based on header + example data (rows 1-3)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:4]:  # rows 1-4
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def apply_dollar_format(ws, dollar_cols, start_row=5, end_row=5000):
    """Apply currency number format to dollar columns."""
    for col_idx in dollar_cols:
        col_letter = get_column_letter(col_idx)
        for row in range(start_row, min(end_row, start_row + 10)):  # format first 10 data rows
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'


def setup_print(ws):
    """Landscape, fit all columns on one page, repeat header rows."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:2'
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)


def add_sheet(name, headers, required_flags, examples, notes=None,
              validations=None, tab_color=TAB_PARTS, dollar_cols=None):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = tab_color
    num_cols = len(headers)

    # Row 1: Headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, num_cols)

    # Row 2: Required/Optional
    style_indicator_row(ws, num_cols, required_flags)

    # Row 3: Example data (yellow tint)
    style_example_row(ws, examples)

    # Row 4: Notes (blue tint, with comments for long values)
    style_notes_row(ws, notes)

    # Data validations (applied to rows 5-5000)
    if validations:
        for col_idx, choices in validations.items():
            if choices:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(choices)}"',
                    allow_blank=True,
                    showErrorMessage=True,
                    showInputMessage=True,
                    errorTitle="Invalid Value",
                    error=f"Must be one of: {', '.join(choices)}",
                    promptTitle="Allowed Values",
                    prompt=", ".join(choices),
                )
                col_letter = get_column_letter(col_idx)
                dv.add(f"{col_letter}5:{col_letter}5000")
                ws.add_data_validation(dv)

    # Dollar formatting
    if dollar_cols:
        apply_dollar_format(ws, dollar_cols)

    # Freeze panes so rows 1-4 stay visible
    ws.freeze_panes = "A5"

    # Auto-width
    auto_width(ws)

    # Print setup
    setup_print(ws)

    return ws


# ════════════════════════════════════════════════════════════════
# SHEET 0: INSTRUCTIONS (default sheet)
# ════════════════════════════════════════════════════════════════
ws_inst = wb.active
ws_inst.title = "README"
ws_inst.sheet_properties.tabColor = TAB_README

instructions = [
    ("ATHELON MRO — PARTS DEPARTMENT BULK UPLOAD TEMPLATE", "", ""),
    ("", "", ""),
    ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
    ("", "", ""),
    ("HOW TO USE THIS TEMPLATE", "", ""),
    ("", "", ""),
    ("1.", "Each tab represents a different data category in the Parts department.", ""),
    ("2.", "Row 1 = Column headers. DO NOT modify these.", ""),
    ("3.", "Row 2 = Required (red) vs Optional (gray) indicator per column.", ""),
    ("4.", "Row 3 = Example data (yellow row). DELETE this row before uploading.", ""),
    ("5.", "Row 4 = Notes & allowed values (blue row). DELETE before uploading.", ""),
    ("6.", "Start entering your data from Row 5 onward.", ""),
    ("7.", "Columns marked REQUIRED must have a value in every data row.", ""),
    ("8.", "Dropdown validations are applied — click a cell to see allowed values.", ""),
    ("9.", "Hover over truncated notes cells to see full allowed-value lists.", ""),
    ("", "", ""),
    ("SHEETS IN THIS WORKBOOK", "", ""),
    ("", "", ""),
    ("Tab", "Sheet Name", "Description"),
]

sheet_descriptions = [
    ("1", "Standard Parts", "Non-serialized standard & expendable parts"),
    ("2", "Serialized Parts", "Parts tracked by individual serial number"),
    ("3", "Consumables", "Consumable items (fluids, hardware, chemicals)"),
    ("4", "Rotables", "Rotable components with TSN/TSO/TBO tracking"),
    ("5", "Repairable Parts", "Repairable components with disposition tracking"),
    ("6", "Lots & Batches", "Lot/batch tracking for consumables & bulk items"),
    ("7", "Tool Crib", "Hand tools, power tools, special tooling"),
    ("8", "Test Equipment", "Calibrated test equipment & instruments"),
    ("9", "Warehouse Setup", "Warehouse / Area / Shelf / Location / Bin hierarchy"),
    ("10", "Vendors & Suppliers", "Approved vendor & supplier records"),
    ("11", "Loaners", "Rental and loaner component inventory"),
    ("12", "Tag Categories", "Tag taxonomy categories"),
    ("13", "Tags", "Individual tags within categories"),
    ("14", "Labor Kits", "Pre-defined labor/parts kits for common tasks"),
    ("15", "Purchase Orders", "Purchase order header + line item records"),
    ("16", "Core Tracking", "Core exchange return tracking records"),
]
instructions.extend(sheet_descriptions)

instructions.extend([
    ("", "", ""),
    ("DATA FORMAT RULES", "", ""),
    ("", "", ""),
    ("Rule", "Details", "Example"),
    ("Part Numbers", "Automatically uppercased on import", "ms20470ad4-6 → MS20470AD4-6"),
    ("Dates", "Use YYYY-MM-DD format", "2026-03-15"),
    ("Dollar amounts", "Plain decimal numbers — no $ sign, no commas", "1250.00"),
    ("Booleans", "Use TRUE or FALSE (all caps)", "TRUE"),
    ("Condition values", "new, serviceable, overhauled, repaired, unserviceable, quarantine, scrapped", "serviceable"),
    ("Part categories", "consumable, standard, rotable, expendable, repairable", "rotable"),
    ("", "", ""),
    ("REGULATORY NOTES", "", ""),
    ("", "", ""),
    ("INV-23", "All received parts start in 'pending_inspection' status automatically.", ""),
    ("INV-07", "8130-3 trace documents are required for owner-supplied parts (OSP).", ""),
    ("INV-11", "Life-limited parts (LLP) enforce hard-block at zero remaining life.", ""),
    ("INV-12", "Shelf-life expiry triggers alerts at 90 / 60 / 30 day windows.", ""),
])

for row_idx, (col_a, col_b, col_c) in enumerate(instructions, 1):
    cell_a = ws_inst.cell(row=row_idx, column=1, value=col_a)
    if col_b:
        ws_inst.cell(row=row_idx, column=2, value=col_b)
    if col_c:
        ws_inst.cell(row=row_idx, column=3, value=col_c)

    # Title
    if row_idx == 1:
        cell_a.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    # Section headers
    elif col_a in ("HOW TO USE THIS TEMPLATE", "SHEETS IN THIS WORKBOOK",
                    "DATA FORMAT RULES", "REGULATORY NOTES"):
        cell_a.font = Font(name="Calibri", bold=True, size=13, color="2E75B6")
    # Table headers
    elif col_a in ("Tab", "Rule"):
        for c in range(1, 4):
            cell = ws_inst.cell(row=row_idx, column=c)
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
    # Sheet list rows
    elif col_a.isdigit() and 1 <= int(col_a) <= 16:
        cell_a.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        cell_a.alignment = Alignment(horizontal="center")
        for c in range(1, 4):
            ws_inst.cell(row=row_idx, column=c).border = THIN_BORDER
    # Numbered steps
    elif col_a.endswith(".") and col_a[:-1].isdigit():
        cell_a.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        cell_a.alignment = Alignment(horizontal="right")
    # Regulatory codes
    elif col_a.startswith("INV-"):
        cell_a.font = Font(name="Calibri", bold=True, size=10, color="CC0000")
        for c in range(1, 4):
            ws_inst.cell(row=row_idx, column=c).border = THIN_BORDER
    # Format rule rows
    elif col_b and col_c and col_a not in ("", "Generated:"):
        cell_a.font = Font(name="Calibri", bold=True, size=10)
        for c in range(1, 4):
            ws_inst.cell(row=row_idx, column=c).border = THIN_BORDER

ws_inst.column_dimensions["A"].width = 22
ws_inst.column_dimensions["B"].width = 65
ws_inst.column_dimensions["C"].width = 40
ws_inst.sheet_view.showGridLines = False


# ════════════════════════════════════════════════════════════════
# SHEET 1: STANDARD PARTS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Standard Parts",
    headers=[
        "Part Number", "Part Name", "Description", "Part Category",
        "Condition", "Qty On Hand", "Min Stock Level", "Reorder Point",
        "Unit Cost ($)", "Supplier", "PO Number",
        "Warehouse Zone", "Aisle", "Shelf", "Bin Number", "Bin Location",
        "Has Shelf Life", "Shelf Life Expiry",
        "Is Life Limited", "Life Limit Hrs", "Life Limit Cyc",
        "Is Owner Supplied", "Lot Number", "Batch Number",
        "Lead Time (Days)", "Notes"
    ],
    required_flags=[
        True, True, False, True,
        True, False, False, False,
        False, False, False,
        False, False, False, False, False,
        True, False,
        True, False, False,
        True, False, False,
        False, False
    ],
    examples=[
        "MS20470AD4-6", "AN Rivet 1/8 x 3/8", "Universal head aluminum rivet", "standard",
        "new", 500, 100, 200,
        0.12, "Aircraft Spruce", "PO-2026-0042",
        "A", "1", "3", "B-14", "A-1-3-B14",
        "FALSE", "",
        "FALSE", "", "",
        "FALSE", "LOT-2026-001", "BATCH-A",
        14, ""
    ],
    notes=[
        "Auto-uppercased", "", "", "standard | expendable",
        "new | serviceable | overhauled | repaired", "Integer", "Alert threshold", "Auto-PO threshold",
        "Decimal, no $", "", "",
        "", "", "", "", "Full path display",
        "TRUE / FALSE", "YYYY-MM-DD if TRUE",
        "TRUE / FALSE", "Hours if TRUE", "Cycles if TRUE",
        "TRUE / FALSE", "", "",
        "Integer", ""
    ],
    validations={
        4: ["standard", "expendable"],
        5: ["new", "serviceable", "overhauled", "repaired", "unserviceable"],
        17: ["TRUE", "FALSE"],
        19: ["TRUE", "FALSE"],
        22: ["TRUE", "FALSE"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[9],
)


# ════════════════════════════════════════════════════════════════
# SHEET 2: SERIALIZED PARTS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Serialized Parts",
    headers=[
        "Part Number", "Part Name", "Description", "Part Category",
        "Serial Number", "Condition",
        "Supplier", "PO Number", "Receiving Date",
        "Is Life Limited", "Life Limit Hrs", "Life Limit Cyc",
        "Hours Accum (TSN)", "Cycles Accum",
        "Has Shelf Life", "Shelf Life Expiry",
        "Is Owner Supplied",
        "Warehouse Zone", "Aisle", "Shelf", "Bin Number",
        "Unit Cost ($)", "Lead Time (Days)", "Notes"
    ],
    required_flags=[
        True, True, False, True,
        True, True,
        False, False, False,
        True, False, False,
        False, False,
        True, False,
        True,
        False, False, False, False,
        False, False, False
    ],
    examples=[
        "SL61398", "Slick Magneto", "Dual magneto, 4-cyl Lycoming", "standard",
        "SN-88421", "serviceable",
        "Tempest Aero", "PO-2026-0055", "2026-03-01",
        "TRUE", 2000, 0,
        1245.6, 843,
        "FALSE", "",
        "FALSE",
        "B", "2", "1", "C-07",
        3450.00, 21, "Overhauled unit, 8130-3 on file"
    ],
    notes=[
        "Auto-uppercased", "", "", "standard | rotable | repairable",
        "Unique per unit", "new | serviceable | overhauled | repaired | unserviceable",
        "", "", "YYYY-MM-DD",
        "TRUE / FALSE", "If life-limited", "If life-limited",
        "Time since new/overhaul", "",
        "TRUE / FALSE", "YYYY-MM-DD",
        "TRUE / FALSE",
        "", "", "", "",
        "Decimal, no $", "Integer", ""
    ],
    validations={
        4: ["standard", "rotable", "repairable"],
        6: ["new", "serviceable", "overhauled", "repaired", "unserviceable"],
        10: ["TRUE", "FALSE"],
        15: ["TRUE", "FALSE"],
        17: ["TRUE", "FALSE"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[22],
)


# ════════════════════════════════════════════════════════════════
# SHEET 3: CONSUMABLES
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Consumables",
    headers=[
        "Part Number", "Part Name", "Description", "Part Category",
        "Condition", "Qty On Hand", "Min Stock Level", "Reorder Point",
        "Unit Cost ($)", "Avg Cost ($)", "Last Purchase ($)",
        "Supplier", "PO Number",
        "Lot Number", "Batch Number",
        "Has Shelf Life", "Shelf Life Expiry",
        "Warehouse Zone", "Aisle", "Shelf", "Bin Number", "Bin Location",
        "Lead Time (Days)", "Notes"
    ],
    required_flags=[
        True, True, False, True,
        True, False, False, False,
        False, False, False,
        False, False,
        False, False,
        True, False,
        False, False, False, False, False,
        False, False
    ],
    examples=[
        "AMS-S-8802", "Aero-Seal Sealant", "Fuel tank sealant, 6oz tube", "consumable",
        "new", 24, 6, 12,
        18.50, 17.85, 19.00,
        "PRC-DeSoto", "PO-2026-0061",
        "LOT-2026-015", "BATCH-C",
        "TRUE", "2027-06-15",
        "C", "1", "2", "H-03", "C-1-2-H03",
        7, "Temperature controlled storage required"
    ],
    notes=[
        "Auto-uppercased", "", "", "Always 'consumable'",
        "new | serviceable", "Integer", "Alert threshold", "Auto-PO threshold",
        "Decimal, no $", "Calculated field", "From last PO",
        "", "",
        "", "",
        "TRUE / FALSE", "YYYY-MM-DD if TRUE",
        "", "", "", "", "Full path",
        "Integer", ""
    ],
    validations={
        4: ["consumable"],
        5: ["new", "serviceable"],
        16: ["TRUE", "FALSE"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[9, 10, 11],
)


# ════════════════════════════════════════════════════════════════
# SHEET 4: ROTABLES
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Rotables",
    headers=[
        "Part Number", "Serial Number", "Description",
        "Status", "Condition",
        "Position Code",
        "TSN Hours", "TSN Cycles", "TSO Hours", "TSO Cycles",
        "TSI Hours", "TSI Cycles", "TBO Hours", "TBO Cycles",
        "Shelf Life Expiry",
        "Purchase Price ($)", "Current Value ($)", "Core Value ($)",
        "Last OH Vendor", "Last OH Date",
        "Warranty Expiry",
        "Notes"
    ],
    required_flags=[
        True, True, True,
        True, True,
        False,
        False, False, False, False,
        False, False, False, False,
        False,
        False, False, False,
        False, False,
        False,
        False
    ],
    examples=[
        "RA-215CC", "SN-CC-77210", "McCauley 3-blade constant speed prop",
        "serviceable", "overhauled",
        "PROP-1",
        4521.3, 3102, 245.8, 167,
        245.8, 167, 2000, 0,
        "",
        28500.00, 22000.00, 8500.00,
        "McCauley Propeller Systems", "2025-11-20",
        "2027-11-20",
        "Core exchange agreement with vendor"
    ],
    notes=[
        "Auto-uppercased", "Unique", "Max 200 chars",
        "installed | serviceable | in_shop | at_vendor | condemned | loaned_out",
        "serviceable | unserviceable | overhauled | repaired | inspected",
        "e.g. PROP-1, ENG-L, ALT-1",
        "Time Since New", "", "Time Since Overhaul", "",
        "Time Since Inspection", "", "TBO limit", "",
        "YYYY-MM-DD",
        "Decimal, no $", "Decimal, no $", "Decimal, no $",
        "", "YYYY-MM-DD",
        "YYYY-MM-DD",
        ""
    ],
    validations={
        4: ["installed", "serviceable", "in_shop", "at_vendor", "condemned", "loaned_out"],
        5: ["serviceable", "unserviceable", "overhauled", "repaired", "inspected"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[16, 17, 18],
)


# ════════════════════════════════════════════════════════════════
# SHEET 5: REPAIRABLE PARTS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Repairable Parts",
    headers=[
        "Part Number", "Part Name", "Description",
        "Serial Number", "Condition",
        "Is Life Limited", "Life Limit Hrs", "Life Limit Cyc",
        "Hours Accum (TSN)", "Cycles Accum",
        "Has Shelf Life", "Shelf Life Expiry",
        "Supplier", "PO Number", "Receiving Date",
        "Unit Cost ($)", "Is Owner Supplied",
        "Warehouse Zone", "Aisle", "Shelf", "Bin Number",
        "Intended Disposition", "Notes"
    ],
    required_flags=[
        True, True, False,
        True, True,
        True, False, False,
        False, False,
        True, False,
        False, False, False,
        False, True,
        False, False, False, False,
        False, False
    ],
    examples=[
        "PN-ALT-3020", "Alternator Assembly", "28V 60A alternator, Cessna 172",
        "SN-ALT-9921", "unserviceable",
        "TRUE", 3000, 0,
        2876.4, 0,
        "FALSE", "",
        "Plane-Power", "PO-2026-0070", "2026-02-15",
        4200.00, "FALSE",
        "D", "1", "4", "R-02",
        "Send to vendor for overhaul", "Removed for intermittent output"
    ],
    notes=[
        "Auto-uppercased", "", "",
        "Unique per unit", "new | serviceable | overhauled | repaired | unserviceable",
        "TRUE / FALSE", "", "",
        "", "",
        "TRUE / FALSE", "YYYY-MM-DD",
        "", "", "YYYY-MM-DD",
        "Decimal, no $", "TRUE / FALSE",
        "", "", "", "",
        "Free text", ""
    ],
    validations={
        5: ["new", "serviceable", "overhauled", "repaired", "unserviceable"],
        6: ["TRUE", "FALSE"],
        11: ["TRUE", "FALSE"],
        17: ["TRUE", "FALSE"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[16],
)


# ════════════════════════════════════════════════════════════════
# SHEET 6: LOTS & BATCHES
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Lots & Batches",
    headers=[
        "Lot Number", "Batch Number", "Part Number", "Part Name", "Description",
        "Original Qty", "Received Qty",
        "Condition",
        "Vendor Name",
        "PO Number",
        "Has Shelf Life", "Shelf Life Expiry",
        "Bin Location Code",
        "Notes"
    ],
    required_flags=[
        False, False, True, True, False,
        True, True,
        False,
        False,
        False,
        True, False,
        False,
        False
    ],
    examples=[
        "LOT-2026-042", "BATCH-B", "AN3-5A", "AN Bolt 10-32 x 9/16", "Cadmium plated steel bolt",
        1000, 1000,
        "new",
        "Aircraft Spruce",
        "PO-2026-0042",
        "FALSE", "",
        "A-1-3-B14",
        "Certificate of Conformity on file"
    ],
    notes=[
        "Auto-generated if blank", "", "Must match Parts", "", "",
        "Total in lot", "Qty received",
        "new | serviceable | quarantine | expired | depleted",
        "Must match Vendors tab",
        "",
        "TRUE / FALSE", "YYYY-MM-DD if TRUE",
        "Warehouse bin code",
        ""
    ],
    validations={
        8: ["new", "serviceable", "quarantine", "expired", "depleted"],
        11: ["TRUE", "FALSE"],
    },
    tab_color=TAB_PARTS,
)


# ════════════════════════════════════════════════════════════════
# SHEET 7: TOOL CRIB
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Tool Crib",
    headers=[
        "Tool Number", "Description", "Serial Number",
        "Category", "Location",
        "Cal Required", "Cal Interval (Days)",
        "Cal Provider",
        "Last Cal Date", "Next Cal Due",
        "Purchase Date",
        "Notes"
    ],
    required_flags=[
        True, True, False,
        True, False,
        True, False,
        False,
        False, False,
        False,
        False
    ],
    examples=[
        "TL-001", "Torque Wrench 1/2\" drive", "SN-TW-44210",
        "hand_tool", "Bay 2, Cabinet A",
        "TRUE", 365,
        "Snap-on Tools Calibration",
        "2025-12-01", "2026-12-01",
        "2023-06-15",
        "25-250 ft-lbs range"
    ],
    notes=[
        "Unique tool ID", "", "",
        "hand_tool | power_tool | test_equipment | special_tooling | consumable",
        "Free text location",
        "TRUE / FALSE", "Days between cals",
        "Cal vendor name",
        "YYYY-MM-DD", "YYYY-MM-DD",
        "YYYY-MM-DD",
        ""
    ],
    validations={
        4: ["hand_tool", "power_tool", "test_equipment", "special_tooling", "consumable"],
        6: ["TRUE", "FALSE"],
    },
    tab_color=TAB_TOOLS,
)


# ════════════════════════════════════════════════════════════════
# SHEET 8: TEST EQUIPMENT
# ════════════════════════════════════════════════════════════════
# Equipment types list (too long for a single cell — use comment)
_equip_types = [
    "multimeter", "oscilloscope", "pitot_static_tester", "altimeter_tester",
    "transponder_tester", "nav_com_tester", "torque_wrench", "borescope",
    "rigging_tool", "pressure_gauge", "fuel_flow_tester", "insulation_tester",
    "timing_light", "compression_tester", "prop_balance_analyzer", "other"
]

add_sheet(
    "Test Equipment",
    headers=[
        "Part Number", "Serial Number", "Equipment Name", "Manufacturer",
        "Equipment Type",
        "Cal Cert Number", "Cal Date", "Cal Expiry Date",
        "Cal Performed By",
        "Status",
        "Notes"
    ],
    required_flags=[
        True, True, True, False,
        True,
        True, True, True,
        False,
        True,
        False
    ],
    examples=[
        "FLUKE-87V", "SN-FL-12345", "Fluke 87V Digital Multimeter", "Fluke",
        "multimeter",
        "CAL-2026-0089", "2026-01-15", "2027-01-15",
        "Precision Calibration Labs",
        "current",
        "Primary bench meter"
    ],
    notes=[
        "", "Unique", "", "",
        "See dropdown (16 types)",
        "", "YYYY-MM-DD", "YYYY-MM-DD",
        "",
        "current | expired | out_for_calibration | removed_from_service | quarantine",
        ""
    ],
    validations={
        5: _equip_types,
        10: ["current", "expired", "out_for_calibration", "removed_from_service", "quarantine"],
    },
    tab_color=TAB_TOOLS,
)


# ════════════════════════════════════════════════════════════════
# SHEET 9: WAREHOUSE SETUP
# ════════════════════════════════════════════════════════════════
ws_wh = add_sheet(
    "Warehouse Setup",
    headers=[
        "Level", "Warehouse Code", "Warehouse Name",
        "Area Code", "Area Name", "Area Type",
        "Shelf Code", "Shelf Name",
        "Location Code", "Location Name",
        "Bin Code", "Bin Name", "Barcode",
        "Description", "Address"
    ],
    required_flags=[
        True, True, True,
        False, False, False,
        False, False,
        False, False,
        False, False, False,
        False, False
    ],
    examples=[
        "warehouse", "WH-MAIN", "Main Warehouse",
        "", "", "",
        "", "",
        "", "",
        "", "", "",
        "Primary parts storage", "123 Hangar Rd, Hangar B"
    ],
    notes=[
        "warehouse | area | shelf | location | bin",
        "Parent warehouse code",
        "Required for warehouse level",
        "Required for area+ levels",
        "Required for area level",
        "general | hazmat | temperature_controlled | secure | quarantine | receiving",
        "Required for shelf+ levels", "",
        "Required for location+ levels", "",
        "Required for bin level", "", "For barcode scanning",
        "", "Warehouse level only"
    ],
    validations={
        1: ["warehouse", "area", "shelf", "location", "bin"],
        6: ["general", "hazmat", "temperature_controlled", "secure", "quarantine", "receiving"],
    },
    tab_color=TAB_WAREHOUSE,
)

# Additional hierarchy examples showing the full tree pattern
hierarchy_examples = [
    # row 5-12: a complete warehouse hierarchy example
    ["area",     "WH-MAIN", "",              "A",  "General Storage", "general",     "",   "",         "",   "",           "",    "",       "",                  "Main floor",            ""],
    ["area",     "WH-MAIN", "",              "HZ", "Hazmat Storage",  "hazmat",      "",   "",         "",   "",           "",    "",       "",                  "Chemical storage area", ""],
    ["area",     "WH-MAIN", "",              "Q",  "Quarantine",      "quarantine",  "",   "",         "",   "",           "",    "",       "",                  "Hold area",             ""],
    ["area",     "WH-MAIN", "",              "R",  "Receiving Dock",  "receiving",   "",   "",         "",   "",           "",    "",       "",                  "Inbound receiving",     ""],
    ["shelf",    "WH-MAIN", "",              "A",  "",                "",            "S1", "Shelf 1",  "",   "",           "",    "",       "",                  "",                      ""],
    ["location", "WH-MAIN", "",              "A",  "",                "",            "S1", "",         "L1", "Location 1", "",    "",       "",                  "",                      ""],
    ["bin",      "WH-MAIN", "",              "A",  "",                "",            "S1", "",         "L1", "",           "B01", "Bin 01", "WH-A-S1-L1-B01",   "",                      ""],
    ["bin",      "WH-MAIN", "",              "A",  "",                "",            "S1", "",         "L1", "",           "B02", "Bin 02", "WH-A-S1-L1-B02",   "",                      ""],
]
for row_idx, row_data in enumerate(hierarchy_examples, 5):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_wh.cell(row=row_idx, column=col_idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER

# Add a separator row after examples
sep_row = len(hierarchy_examples) + 5
for col in range(1, 16):
    cell = ws_wh.cell(row=sep_row, column=col)
    cell.value = "▼ DELETE EXAMPLES ABOVE — ENTER YOUR DATA BELOW ▼" if col == 1 else ""
    cell.font = Font(name="Calibri", bold=True, size=10, color="CC0000")
    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    cell.border = THIN_BORDER
ws_wh.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=15)
ws_wh.cell(row=sep_row, column=1).alignment = Alignment(horizontal="center")


# ════════════════════════════════════════════════════════════════
# SHEET 10: VENDORS & SUPPLIERS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Vendors & Suppliers",
    headers=[
        "Vendor Name", "Type",
        "Contact Name", "Contact Email", "Contact Phone",
        "Address",
        "Certificate Number", "Cert Expiry Date",
        "Is Approved", "Approved By",
        "Default Lead Time (Days)",
        "Notes"
    ],
    required_flags=[
        True, True,
        False, False, False,
        False,
        False, False,
        True, False,
        False,
        False
    ],
    examples=[
        "Aircraft Spruce & Specialty", "parts_supplier",
        "John Smith", "john@aircraftspruce.com", "951-372-9555",
        "225 Airport Circle, Corona, CA 92880",
        "AS-2345-SUP", "2027-12-31",
        "TRUE", "Jane Doe",
        14,
        "Primary hardware supplier"
    ],
    notes=[
        "", "parts_supplier | consumables_supplier | contract_maintenance | calibration_lab | DER | service_provider | other",
        "", "", "",
        "",
        "FAR 145 or supplier cert", "YYYY-MM-DD",
        "TRUE / FALSE", "Approver name",
        "Integer",
        ""
    ],
    validations={
        2: ["parts_supplier", "consumables_supplier", "contract_maintenance", "calibration_lab", "DER", "service_provider", "other"],
        9: ["TRUE", "FALSE"],
    },
    tab_color=TAB_VENDORS,
)


# ════════════════════════════════════════════════════════════════
# SHEET 11: LOANERS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Loaners",
    headers=[
        "Part Number", "Serial Number", "Description",
        "Status",
        "Daily Rate ($)",
        "Notes"
    ],
    required_flags=[
        True, False, True,
        True,
        False,
        False
    ],
    examples=[
        "PN-GEN-500", "SN-GEN-001", "Portable GPU 28V DC",
        "available",
        150.00,
        "Includes cable set and cart"
    ],
    notes=[
        "Auto-uppercased", "", "Max 200 chars",
        "available | loaned_out | maintenance | retired",
        "Per-day rental rate",
        ""
    ],
    validations={
        4: ["available", "loaned_out", "maintenance", "retired"],
    },
    tab_color=TAB_PARTS,
    dollar_cols=[5],
)


# ════════════════════════════════════════════════════════════════
# SHEET 12: TAG CATEGORIES
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Tag Categories",
    headers=[
        "Category Name", "Slug", "Category Type",
        "Description", "Display Order", "Is System", "Is Active"
    ],
    required_flags=[
        True, True, True,
        False, True, True, True
    ],
    examples=[
        "Cessna Models", "cessna-models", "aircraft_type",
        "Tags for Cessna aircraft compatibility", 1, "FALSE", "TRUE"
    ],
    notes=[
        "", "lowercase-dashes", "aircraft_type | engine_type | ata_chapter | component_type | custom",
        "", "Sort order (integer)", "TRUE = system-seeded", "TRUE / FALSE"
    ],
    validations={
        3: ["aircraft_type", "engine_type", "ata_chapter", "component_type", "custom"],
        6: ["TRUE", "FALSE"],
        7: ["TRUE", "FALSE"],
    },
    tab_color=TAB_CONFIG,
)


# ════════════════════════════════════════════════════════════════
# SHEET 13: TAGS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Tags",
    headers=[
        "Category Slug", "Tag Name", "Tag Code",
        "Description",
        "Aircraft Make", "Aircraft Model",
        "Engine Make", "Engine Model",
        "Display Order", "Is Active"
    ],
    required_flags=[
        True, True, False,
        False,
        False, False,
        False, False,
        True, True
    ],
    examples=[
        "cessna-models", "Cessna 172S", "C172S",
        "Cessna 172S Skyhawk SP",
        "Cessna", "172S",
        "", "",
        1, "TRUE"
    ],
    notes=[
        "Must match Tag Categories slug", "", "Short code",
        "",
        "For aircraft_type tags", "For aircraft_type tags",
        "For engine_type tags", "For engine_type tags",
        "Sort order (integer)", "TRUE / FALSE"
    ],
    validations={
        10: ["TRUE", "FALSE"],
    },
    tab_color=TAB_CONFIG,
)


# ════════════════════════════════════════════════════════════════
# SHEET 14: LABOR KITS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Labor Kits",
    headers=[
        "Kit Name", "Description", "ATA Chapter",
        "Aircraft Type", "Estimated Hrs", "Labor Rate ($/hr)",
        "Part 1 P/N", "Part 1 Desc", "Part 1 Qty", "Part 1 Cost",
        "Part 2 P/N", "Part 2 Desc", "Part 2 Qty", "Part 2 Cost",
        "Part 3 P/N", "Part 3 Desc", "Part 3 Qty", "Part 3 Cost",
        "Part 4 P/N", "Part 4 Desc", "Part 4 Qty", "Part 4 Cost",
        "Part 5 P/N", "Part 5 Desc", "Part 5 Qty", "Part 5 Cost",
        "Is Active"
    ],
    required_flags=[
        True, False, False,
        False, True, False,
        False, False, False, False,
        False, False, False, False,
        False, False, False, False,
        False, False, False, False,
        False, False, False, False,
        True
    ],
    examples=[
        "Annual Inspection - C172", "100-hr / Annual kit", "05",
        "Cessna 172", 24.0, 95.00,
        "CH48108-1", "Oil Filter", 1, 18.50,
        "AMS-S-8802", "Gasket Sealant", 1, 12.00,
        "MS20470AD4-6", "AN Rivets", 50, 0.12,
        "", "", "", "",
        "", "", "", "",
        "TRUE"
    ],
    notes=[
        "", "", "ATA chapter code",
        "", "Total est. hours", "$/hr",
        "Part number", "", "Integer", "Unit $",
        "", "", "", "",
        "", "", "", "",
        "Add columns for >5 parts", "", "", "",
        "Add columns for >5 parts", "", "", "",
        "TRUE / FALSE"
    ],
    validations={
        27: ["TRUE", "FALSE"],
    },
    tab_color=TAB_CONFIG,
    dollar_cols=[6, 10, 14, 18, 22, 26],
)


# ════════════════════════════════════════════════════════════════
# SHEET 15: PURCHASE ORDERS
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Purchase Orders",
    headers=[
        "PO Number", "Vendor Name", "Status",
        "Currency",
        "Line 1 P/N", "Line 1 Desc", "Line 1 Qty", "Line 1 Price",
        "Line 2 P/N", "Line 2 Desc", "Line 2 Qty", "Line 2 Price",
        "Line 3 P/N", "Line 3 Desc", "Line 3 Qty", "Line 3 Price",
        "Line 4 P/N", "Line 4 Desc", "Line 4 Qty", "Line 4 Price",
        "Line 5 P/N", "Line 5 Desc", "Line 5 Qty", "Line 5 Price",
        "Notes"
    ],
    required_flags=[
        True, True, True,
        False,
        True, True, True, True,
        False, False, False, False,
        False, False, False, False,
        False, False, False, False,
        False, False, False, False,
        False
    ],
    examples=[
        "PO-2026-0042", "Aircraft Spruce", "DRAFT",
        "USD",
        "MS20470AD4-6", "AN Rivet 1/8 x 3/8", 500, 0.12,
        "CH48108-1", "Champion Oil Filter", 6, 18.50,
        "", "", "", "",
        "", "", "", "",
        "", "", "", "",
        "Rush order for AOG C172"
    ],
    notes=[
        "Unique PO ID", "Must match Vendors tab", "DRAFT | SUBMITTED | PARTIAL | RECEIVED | CLOSED",
        "e.g. USD",
        "", "", "Integer", "Unit price",
        "", "", "", "",
        "", "", "", "",
        "", "", "", "",
        "", "", "", "",
        ""
    ],
    validations={
        3: ["DRAFT", "SUBMITTED", "PARTIAL", "RECEIVED", "CLOSED"],
    },
    tab_color=TAB_VENDORS,
    dollar_cols=[8, 12, 16, 20, 24],
)


# ════════════════════════════════════════════════════════════════
# SHEET 16: CORE TRACKING
# ════════════════════════════════════════════════════════════════
add_sheet(
    "Core Tracking",
    headers=[
        "Core Number", "Part Number", "Serial Number",
        "Description", "Status",
        "Vendor Name",
        "Core Value ($)", "Credit Amount ($)",
        "Return Due Date",
        "Notes"
    ],
    required_flags=[
        False, True, False,
        True, True,
        False,
        True, False,
        False,
        False
    ],
    examples=[
        "CORE-2026-001", "RA-215CC", "SN-CC-77210",
        "McCauley prop - core return", "awaiting_return",
        "McCauley Propeller Systems",
        8500.00, "",
        "2026-04-15",
        "Exchange agreement per PO-2026-0055"
    ],
    notes=[
        "Auto-generated if blank", "Auto-uppercased", "",
        "", "awaiting_return | received | inspected | credit_issued | scrapped | overdue",
        "Must match Vendors tab",
        "Core charge amount", "Credit after inspection",
        "YYYY-MM-DD",
        ""
    ],
    validations={
        5: ["awaiting_return", "received", "inspected", "credit_issued", "scrapped", "overdue"],
    },
    tab_color=TAB_VENDORS,
    dollar_cols=[7, 8],
)


# ════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════
output_path = "/Users/samuelsandifer/conductor/workspaces/Athalon-Open-Claw/nagoya/apps/athelon-app/docs/ops/Athelon_Parts_Bulk_Upload_Template.xlsx"
wb.save(output_path)
print(f"✓ Saved to: {output_path}")
print(f"  Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")
